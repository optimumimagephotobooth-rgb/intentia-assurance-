"""
Intentia Assurance — Operator Console Generator
Creates a single spreadsheet with 6 tabs as specified.
Exportable as .xlsx and compatible with Google Sheets.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_operator_console():
    """Create the Intentia Assurance Operator Console spreadsheet."""
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ===== TAB 1: CASE_REGISTER =====
    ws1 = wb.create_sheet("CASE_REGISTER")
    
    # Headers
    headers = [
        "Case ID",
        "Client / Organisation",
        "Use Case",
        "Intake Date",
        "Status",
        "Freeze Confirmed (Y/N)",
        "Final PDF Sent Date",
        "Operator Notes (Private)",
        "Operator Guidance (System)",
        "Days Since Intake",
        "Days Since Freeze"
    ]
    
    ws1.append(headers)
    
    # Format headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add formulas for Days Since columns
    # Formula for Days Since Intake: =IF(D2<>"",TODAY()-D2,"")
    # Formula for Days Since Freeze: =IF(AND(F2="Y",D2<>""),TODAY()-D2,"")
    # These will be added when rows are populated
    
    # Add data validation for Status dropdown
    from openpyxl.worksheet.datavalidation import DataValidation
    status_dv = DataValidation(type="list", formula1='"Intake,Frozen,Delivered,Closed"', allow_blank=True)
    status_dv.error = "Invalid status"
    ws1.add_data_validation(status_dv)
    status_dv.add("E2:E1000")  # Apply to column E
    
    # Set column widths
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 18
    ws1.column_dimensions['G'].width = 18
    ws1.column_dimensions['H'].width = 40
    ws1.column_dimensions['I'].width = 40
    ws1.column_dimensions['J'].width = 18
    ws1.column_dimensions['K'].width = 18
    
    # ===== TAB 2: OPS_DASHBOARD =====
    ws2 = wb.create_sheet("OPS_DASHBOARD")
    
    # Section headers
    sections = [
        "CASES REQUIRING ACTION",
        "",
        "WAITING ON FREEZE CONFIRMATION",
        "",
        "DELIVERY QUEUE",
        "",
        "RECENTLY DELIVERED"
    ]
    
    for section in sections:
        ws2.append([section])
    
    # Format section headers
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(bold=True, size=12)
    
    for row_num in [1, 3, 5, 7]:
        cell = ws2.cell(row=row_num, column=1)
        cell.fill = section_fill
        cell.font = section_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Add note about read-only intent
    ws2.append([""])
    ws2.append(["Note: This tab pulls data from CASE_REGISTER using FILTER/QUERY formulas."])
    ws2.append(["Never change Status automatically. Never suggest decisions. Never infer meaning."])
    
    ws2.column_dimensions['A'].width = 50
    
    # ===== TAB 3: PRESS_RULES =====
    ws3 = wb.create_sheet("PRESS_RULES")
    
    rules = [
        "The press sequence is fixed: Intake → Freeze → Output.",
        "Thinking is permitted only during Intake.",
        "Once Freeze is confirmed, no edits are permitted.",
        "Corrections require a new record.",
        "The PDF is the authoritative artefact.",
        "Sheets, forms, and emails are not records.",
        "We record intent and scope, not outcomes.",
        "We do not assess effectiveness, competence, or impact.",
        "Automation must not create or modify meaning.",
        "If unsure, stop and do not proceed."
    ]
    
    for rule in rules:
        ws3.append([rule])
    
    # Format rules
    for row in range(1, len(rules) + 1):
        cell = ws3.cell(row=row, column=1)
        cell.font = Font(size=11)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    ws3.column_dimensions['A'].width = 80
    
    # ===== TAB 4: USE_CASE_BOUNDARIES =====
    ws4 = wb.create_sheet("USE_CASE_BOUNDARIES")
    
    # Headers
    headers = [
        "Use Case",
        "What This Includes",
        "What This Explicitly Excludes",
        "Immediate Refusal Triggers"
    ]
    
    ws4.append(headers)
    
    # Format headers
    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    data = [
        [
            "Training Assurance",
            "Intent as stated, scope, delivery facts",
            "Outcomes, effectiveness, competence, evaluation",
            "Asked \"Did it work?\" or \"What impact did this have?\""
        ],
        [
            "Programme Decision Record",
            "Decision context at time made",
            "Judgement, hindsight, approval or rejection",
            "Asked \"Was this the right decision?\""
        ],
        [
            "Strategic Intent / Evidence Pack",
            "Declared intent and exclusions",
            "Performance guarantees, accountability transfer",
            "Asked to certify success or readiness"
        ]
    ]
    
    for row in data:
        ws4.append(row)
    
    # Set column widths
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 40
    ws4.column_dimensions['C'].width = 50
    ws4.column_dimensions['D'].width = 50
    
    # Format data rows
    for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # ===== TAB 5: OPERATOR_PLAYBOOK =====
    ws5 = wb.create_sheet("OPERATOR_PLAYBOOK")
    
    # Headers
    headers = [
        "Situation",
        "What To Say (Verbatim)",
        "What NOT To Say"
    ]
    
    ws5.append(headers)
    
    # Format headers
    for cell in ws5[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    data = [
        [
            "Client asks for opinion",
            "\"We do not assess or advise.\"",
            "Any opinion or explanation"
        ],
        [
            "Client asks to edit after freeze",
            "\"That would require a new record.\"",
            "\"I'll just fix it.\""
        ],
        [
            "Client asks if programme worked",
            "\"This record does not assess outcomes.\"",
            "Any interpretation"
        ],
        [
            "Client pressures for speed",
            "\"The sequence protects the record.\"",
            "Apologies or justifications"
        ]
    ]
    
    for row in data:
        ws5.append(row)
    
    # Set column widths
    ws5.column_dimensions['A'].width = 30
    ws5.column_dimensions['B'].width = 40
    ws5.column_dimensions['C'].width = 40
    
    # Format data rows
    for row in ws5.iter_rows(min_row=2, max_row=ws5.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # ===== TAB 6: DAILY_OPERATOR_FLOW =====
    ws6 = wb.create_sheet("DAILY_OPERATOR_FLOW")
    
    # Headers
    headers = [
        "Step Order",
        "Action",
        "Reminder"
    ]
    
    ws6.append(headers)
    
    # Format headers
    for cell in ws6[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    data = [
        [
            "1",
            "Open Ops Dashboard",
            "Do not open documents yet"
        ],
        [
            "2",
            "Review Intake cases",
            "Thinking allowed only here"
        ],
        [
            "3",
            "Chase freeze confirmations",
            "No assembly before freeze"
        ],
        [
            "4",
            "Assemble ONE PDF",
            "No multitasking"
        ],
        [
            "5",
            "Hash and archive",
            "Mechanical step only"
        ],
        [
            "6",
            "Update Case Register",
            "Status update comes last"
        ]
    ]
    
    for row in data:
        ws6.append(row)
    
    # Set column widths
    ws6.column_dimensions['A'].width = 12
    ws6.column_dimensions['B'].width = 30
    ws6.column_dimensions['C'].width = 30
    
    # Format data rows
    for row in ws6.iter_rows(min_row=2, max_row=ws6.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Save workbook
    filename = "Intentia Assurance — Operator Console.xlsx"
    wb.save(filename)
    print(f"Spreadsheet created: {filename}")
    return filename

if __name__ == "__main__":
    create_operator_console()

